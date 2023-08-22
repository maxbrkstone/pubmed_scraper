import pandas as pd
from Bio import Entrez
from openpyxl import Workbook
from tqdm import tqdm
import sys
from http.client import HTTPException

def search_pubmed(keywords=None, journal=None, terms=None):
    if  terms:
        max_terms = terms
    else:
        max_terms = 10000

    articles = []

    if keywords:
        for keyword in keywords:
            query_keyword = ["(" + keyword + "[Text Word])"]
            print("\nSearching keyword", keyword, "...")

            # Construct the query for PubMed
            if journal:
                query = f'{query_keyword} AND ({journal}[Journal]) AND (("clinical trial"[Publication Type]) OR ("journal article"[Publication Type]) OR ("clinical study"[Publication Type])) NOT ("comment"[Publication Type])'
            else:
                query = f'{query_keyword} AND (("clinical trial"[Publication Type]) OR ("journal article"[Publication Type]) OR ("clinical study"[Publication Type])) NOT ("comment"[Publication Type])'

            # Try to perform the search
            try:
                handle = Entrez.esearch(db="pubmed", term=query, retmax=max_terms, datetype='pdat', mindate='2000/01/01', maxdate='2023/01/01') #FIXME: dates not working
                record = Entrez.read(handle)
            except HTTPException:
                handle.close()
                continue
            handle.close()

            # Fetch the article details
            for pmid in tqdm(record["IdList"], desc="Fetching articles"):
                article = {"PMID": pmid}

                # Try to fetch the article details
                try:
                    fetch_handle = Entrez.efetch(db="pubmed", id=pmid, retmode="xml")
                    fetch_result = Entrez.read(fetch_handle)
                except HTTPException:
                    fetch_handle.close()
                    continue
                fetch_handle.close()
                
                if len(fetch_result['PubmedArticle']) > 0: #Journal article, exclude book articles
                    fetch_record = fetch_result['PubmedArticle'][0]
                    
                    if 'AuthorList' in fetch_record['MedlineCitation']['Article']:
                        if all('LastName' in author and 'ForeName' in author for author in fetch_record['MedlineCitation']['Article']['AuthorList']):
                            article["FirstAuthor"] = {'LastName' : fetch_record['MedlineCitation']['Article']['AuthorList'][0]['LastName'],
                                                    'FirstName' : fetch_record['MedlineCitation']['Article']['AuthorList'][0]['ForeName']}
                        
                            article["LastAuthor"] = {'LastName' : fetch_record['MedlineCitation']['Article']['AuthorList'][-1]['LastName'],
                                                    'FirstName' : fetch_record['MedlineCitation']['Article']['AuthorList'][-1]['ForeName']}
                            
                            if 'Year' in fetch_record['MedlineCitation']['Article']['Journal']['JournalIssue']['PubDate']:
                                article["Year"] = fetch_record['MedlineCitation']['Article']['Journal']['JournalIssue']['PubDate']['Year']
                            elif 'MedlineDate' in fetch_record['MedlineCitation']['Article']['Journal']['JournalIssue']['PubDate']:
                                article["Year"] = fetch_record['MedlineCitation']['Article']['Journal']['JournalIssue']['PubDate']['MedlineDate'][0:4]
                            elif(fetch_record['MedlineCitation']['Article']['ArticleDate']):
                                article["Year"] = fetch_record['MedlineCitation']['Article']['ArticleDate'][0]['Year']
                            else:
                                continue

                            #add matched keyword to article
                            article["Keyword"] = keyword

                            articles.append(article)

    elif journal:
        # Perform the search
        query = f'({journal}[Journal]) AND (("clinical trial"[Publication Type]) OR ("journal article"[Publication Type]) OR ("clinical study"[Publication Type])) NOT ("comment"[Publication Type])'

        handle = Entrez.esearch(db="pubmed", term=query, retmax=max_terms, datetype='pdat', mindate='2000/01/01', maxdate='2023/01/01') #FIXME: dates not working
        record = Entrez.read(handle)
        handle.close()

        # Fetch the article details
        for pmid in tqdm(record["IdList"], desc="Fetching articles"):
            article = {"PMID": pmid}

            # Fetch the article details
            fetch_handle = Entrez.efetch(db="pubmed", id=pmid, retmode="xml")
            fetch_result = Entrez.read(fetch_handle)
            fetch_handle.close()
            
            if len(fetch_result['PubmedArticle']) > 0: #Journal article, exclude book articles
                fetch_record = fetch_result['PubmedArticle'][0]
                
                if 'AuthorList' in fetch_record['MedlineCitation']['Article']:
                    if all('LastName' in author and 'ForeName' in author for author in fetch_record['MedlineCitation']['Article']['AuthorList']):
                        article["FirstAuthor"] = {'LastName' : fetch_record['MedlineCitation']['Article']['AuthorList'][0]['LastName'],
                                                    'FirstName' : fetch_record['MedlineCitation']['Article']['AuthorList'][0]['ForeName']}
                        
                        article["LastAuthor"] = {'LastName' : fetch_record['MedlineCitation']['Article']['AuthorList'][-1]['LastName'],
                                                    'FirstName' : fetch_record['MedlineCitation']['Article']['AuthorList'][-1]['ForeName']}

                        if 'Year' in fetch_record['MedlineCitation']['Article']['Journal']['JournalIssue']['PubDate']:
                            article["Year"] = fetch_record['MedlineCitation']['Article']['Journal']['JournalIssue']['PubDate']['Year']
                        elif 'MedlineDate' in fetch_record['MedlineCitation']['Article']['Journal']['JournalIssue']['PubDate']:
                            article["Year"] = fetch_record['MedlineCitation']['Article']['Journal']['JournalIssue']['PubDate']['MedlineDate'][0:4]
                        elif(fetch_record['MedlineCitation']['Article']['ArticleDate']):
                            article["Year"] = fetch_record['MedlineCitation']['Article']['ArticleDate'][0]['Year']
                        else:
                            continue

                        articles.append(article)
    
    return articles

def export_to_excel(keywords, articles, filename):
   # Create an Excel workbook
    wb = Workbook()

    if keywords:
        # Create a dictionary to store articles for each keyword
        keyword_articles = {}
        for article in articles:
            keyword = article["Keyword"]
            if keyword not in keyword_articles:
                # Create a new sheet for the keyword
                if len(keyword) > 31:
                    ws = wb.create_sheet(title=keyword[0:31])
                else:
                    ws = wb.create_sheet(title=keyword[0:31])
                # Write the header row
                ws.append(["Year", "1st Author First Name", "1st Author Last Name", "Last Author First Name", "Last Author Last Name", "PMID"])
                keyword_articles[keyword] = ws
            else:
                ws = keyword_articles[keyword]

            first_author_fn = article["FirstAuthor"]["FirstName"]
            first_author_ln = article["FirstAuthor"]["LastName"]
            last_author_fn = article["LastAuthor"]["FirstName"]
            last_author_ln = article["LastAuthor"]["LastName"]

            # Write the article details to the corresponding sheet
            ws.append([
                article["Year"],
                first_author_fn,
                first_author_ln,
                last_author_fn,
                last_author_ln,
                article["PMID"]
            ])

        # Remove the default sheet created by openpyxl
        del wb["Sheet"]
    else:
        ws = wb.active

        # Write the header row
        ws.append(["Year", "1st Author First Name", "1st Author Last Name", "Last Author First Name", "Last Author Last Name", "PMID"])

        # Write the article details
        for article in articles:
            first_author_fn = article["FirstAuthor"]["FirstName"]
            first_author_ln = article["FirstAuthor"]["LastName"]
            last_author_fn = article["LastAuthor"]["FirstName"]
            last_author_ln = article["LastAuthor"]["LastName"]

            # Write the article details to the corresponding sheet
            ws.append([
                article["Year"],
                first_author_fn,
                first_author_ln,
                last_author_fn,
                last_author_ln,
                article["PMID"]
            ])

    # Save the workbook
    if filename.isspace() or filename == "":
        filename = "articles"

    wb.save(filename + ".xlsx")
    print("Articles exported to " + filename + ".xlsx")

def main():
    # Set your email address for Entrez
    Entrez.email = input("Email to use for PubMed search query: ") 
    # Prompt the user for search criteria
    keywords = input("Enter keywords (comma-separated): ")
    journal = input("Enter journal name (optional): ")
    terms = input("Enter search term limit (optional): ")
    filename = input("Enter filename for excel file ('articles' by default): ") 
    
    if not keywords and not journal:
        print("Must provide at least one keyword or a journal name to search.")
        sys.exit(1)

    if keywords:
        # Split the keywords into a formatted list
        keywords_list = [kw.strip() for kw in keywords.split(", ")]
        #Search PubMed
        articles = search_pubmed(keywords_list, journal, terms)
        # Export articles to Excel
        export_to_excel(keywords_list, articles, filename)
    else:
        articles = search_pubmed(None, journal, terms)
        # Export articles to Excel
        export_to_excel(None, articles, filename)
    

if __name__ == "__main__":
    main()
